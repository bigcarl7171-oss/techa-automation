# -*- coding: utf-8 -*-
"""build-products.py의 LINES/EXTRA_LINES에 박혀 있던 추천 태그를
docs/gift-finder-tags.xlsx로 내보낸다 (일회성 마이그레이션 + 이후 재수출용).

이 파일을 만든 뒤에는 build-products.py가 태그를 코드가 아니라
이 xlsx에서 읽어오도록 바뀐다 — 태그를 고칠 땐 이 xlsx만 고치면 된다.
"""
import sys, os
import importlib.util
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
spec = importlib.util.spec_from_file_location("build_products", os.path.join(REPO_ROOT, 'scripts', 'build-products.py'))
bp = importlib.util.module_from_spec(spec)
spec.loader.exec_module(bp)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

OUT = os.path.join(REPO_ROOT, 'docs', 'gift-finder-tags.xlsx')

HEADERS = ["라인ID", "상품라인명", "분류", "추천대상(recipient)", "상황·목적(occasion)",
           "선물유형(giftType)", "추천상황 설명(situation)", "추천 이유(reason)", "비고(note)"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "라인별 추천태그"

for col, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="8A6D3B")
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

rows = []
for ln in bp.LINES:
    rows.append([ln['id'], ln['name'], ln['cat'],
                 ", ".join(ln['recipient']), ", ".join(ln['occasion']),
                 ", ".join(ln['giftType']), ln['situation'], ln['reason'], ""])
for ln in bp.EXTRA_LINES:
    rows.append([ln['id'], ln['name'], ln['cat'],
                 ", ".join(ln['recipient']), ", ".join(ln['occasion']),
                 ", ".join(ln['giftType']), ln['situation'], ln['reason'], ln.get('note', '')])

for r, row in enumerate(rows, start=2):
    for col, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

widths = [22, 20, 10, 20, 34, 20, 46, 50, 30]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

# 참고용 안내 시트
note_ws = wb.create_sheet("읽어주세요")
note_lines = [
    "이 파일은 /ko/gift-finder/ 페이지가 상품을 추천할 때 쓰는 태그의 편집용 원본입니다.",
    "",
    "- 라인ID / 상품라인명 / 분류 는 손대지 마세요 (가격리스트 xlsx와 매칭하는 키입니다).",
    "- 추천대상·상황·목적·선물유형 은 쉼표(,)로 여러 값을 나열합니다. 예: 연인, 배우자",
    "- 값을 새로 추가하면 사이트 필터에도 자동으로 새 옵션이 생깁니다 (코드 수정 불필요).",
    "- 추천상황 설명 / 추천 이유 는 사이트 카드에 그대로 노출되는 문장입니다.",
    "- 다 고치신 뒤 파일을 저장하고 알려주시면 techa-products.json을 다시 만들어 반영합니다.",
]
for i, t in enumerate(note_lines, start=1):
    note_ws.cell(row=i, column=1, value=t)
note_ws.column_dimensions["A"].width = 100

wb.save(OUT)
print(f"저장: {OUT}  ({len(rows)}개 라인)")
