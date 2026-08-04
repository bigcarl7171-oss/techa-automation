# -*- coding: utf-8 -*-
"""테차상품가격리스트.xlsx -> techa-products.json 생성기.

147개 SKU를 제품 라인으로 묶고, 라인 단위로 다축 태그를 붙인다.
색상/사이즈/전원방식 같은 변형은 라인 아래 variants로 매단다.
"""
import openpyxl, json, io, sys, re, os
from collections import OrderedDict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 이 스크립트(scripts/) 기준 상대경로 — 어느 컴퓨터에서 어느 위치로 clone해도 그대로 동작
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO_ROOT, 'docs', '테차상품가격리스트.xlsx')
OUT = os.path.join(REPO_ROOT, 'assets', 'data', 'techa-products.json')

# ---------------------------------------------------------------------------
# 제품 라인 정의
#   match: SKU 이름이 이 조건을 만족하면 해당 라인에 속함 (순서대로 평가, 먼저 맞는 라인 승)
#   all=반드시 포함할 토큰 / none=포함되면 안 되는 토큰
# ---------------------------------------------------------------------------
LINES = [
    dict(id="rose-hydrangea-bouquet", name="장미수국 꽃다발", cat="꽃다발",
         match=dict(all=["장미수국", "꽃다발"]),
         recipient=["연인", "배우자"],
         occasion=["기념일", "프러포즈", "생일", "화이트데이", "발렌타인데이",
                   "크리스마스", "병문안"],
         giftType=["휴대전달형"],
         situation="연인·배우자에게 기념일이나 프러포즈처럼 마음을 제대로 전하고 싶을 때",
         reason="장미와 수국을 함께 엮은 프리저브드 꽃다발이라 생화처럼 며칠 만에 시들지 않고, 받은 날 그대로의 모습이 오래 남아요."),

    dict(id="hydrangea-bouquet", name="수국 꽃다발", cat="꽃다발",
         match=dict(all=["수국", "꽃다발"], none=["장미수국"]),
         recipient=["연인", "친구", "부모님"], occasion=["생일", "기념일", "병문안"],
         giftType=["휴대전달형"],
         situation="연인·친구·부모님 누구에게 드려도 부담 없는 무난한 꽃 선물을 찾을 때, 또는 물·흙 없는 꽃다발이 필요한 병문안 자리",
         reason="수국은 색이 은은해서 취향을 크게 타지 않고, 프리저브드라 오래 두고 볼 수 있어요. 순수 프리저브드 소재라 물 관리가 필요 없어 병문안에도 부담 없어요."),

    dict(id="superior-rose", name="수페리어 장미", cat="꽃다발",
         match=dict(all=["수페리어"]),
         recipient=["연인", "배우자"],
         occasion=["프러포즈", "기념일", "화이트데이", "발렌타인데이", "병문안"],
         giftType=["휴대전달형"],
         situation="프러포즈나 중요한 기념일처럼 각 잡고 준비하는 자리",
         reason="큰 송이 프리저브드 장미 한 종으로 구성해 꽃 자체가 주인공이 되는 구성이에요."),

    dict(id="rose-soap-bouquet", name="장미 비누꽃다발", cat="꽃다발",
         match=dict(all=["장미", "비누꽃다발"]),
         recipient=["연인", "배우자", "친구"], occasion=["생일", "기념일", "화이트데이", "발렌타인데이"],
         giftType=["휴대전달형"],
         situation="생일·기념일에 향까지 함께 전하고 싶을 때",
         reason="비누로 만든 꽃송이라 은은한 향이 함께 나요. 향이 부담스러우면 무향 옵션도 있습니다."),

    dict(id="pinkpeach-bouquet", name="핑크피치 꽃다발", cat="꽃다발",
         match=dict(all=["핑크피치"]),
         recipient=["연인", "친구"], occasion=["생일", "기념일", "화이트데이"],
         giftType=["휴대전달형"],
         situation="화사한 색감의 꽃다발로 생일을 축하하고 싶을 때",
         reason="핑크와 피치 톤을 섞어 사진에 특히 예쁘게 담기는 조합이에요."),

    dict(id="sunflower-bouquet", name="해바라기 꽃다발", cat="꽃다발",
         match=dict(all=["해바라기", "꽃다발"]),
         recipient=["부모님", "친구", "직장동료"], occasion=["집들이·개업", "승진·퇴임"],
         giftType=["휴대전달형"],
         situation="집들이·개업·승진처럼 좋은 기운을 빌어주고 싶은 자리",
         reason="해바라기는 예로부터 재물운을 상징해 개업·집들이 선물로 많이 찾으세요."),

    dict(id="carnation-rose-bouquet", name="카네이션장미 꽃다발", cat="꽃다발",
         match=dict(all=["카네이션장미", "꽃다발"]),
         recipient=["부모님", "조부모님", "선생님"], occasion=["어버이날", "스승의날", "생신·환갑"],
         giftType=["휴대전달형"],
         situation="어버이날·스승의날에 카네이션을 챙겨야 할 때",
         reason="카네이션과 장미를 함께 엮어 카네이션만 있을 때보다 풍성하고, 시들지 않아 오래 두고 보실 수 있어요."),

    dict(id="carnation-money-bouquet", name="카네이션 돈꽃다발", cat="돈꽃다발",
         match=dict(all=["카네이션", "돈꽃다발"]),
         recipient=["부모님", "조부모님"], occasion=["어버이날", "생신·환갑", "칠순·팔순"],
         giftType=["현금동봉형", "휴대전달형"],
         situation="어버이날이나 생신에 현금을 드리고 싶은데 봉투만 드리기는 밋밋할 때",
         reason="카네이션 사이에 지폐를 꽂는 구조라 현금의 실용성과 꽃의 정성을 같이 전할 수 있어요."),

    dict(id="rose-money-bouquet", name="장미 돈꽃다발", cat="돈꽃다발",
         match=dict(all=["장미", "돈꽃다발"]),
         recipient=["부모님", "조부모님"], occasion=["생신·환갑", "칠순·팔순", "어버이날"],
         giftType=["현금동봉형", "휴대전달형"],
         situation="부모님 생신·환갑·칠순에 현금 선물을 준비할 때",
         reason="며느리·사위·손주가 어른께 드리는 선물로 가장 많이 찾는 구성이에요."),

    dict(id="single-stem-soap-box", name="비누꽃 한송이·두송이 박스", cat="소품",
         match=dict(any_of=["장미한송이", "장미두송이", "카네이션한송이", "카네이션두송이"]),
         recipient=["친구", "직장동료", "선생님", "연인"],
         occasion=["스승의날", "소소한 감사", "단체·답례품",
                   "화이트데이", "발렌타인데이", "빼빼로데이"],
         giftType=["휴대전달형", "단체·답례품"],
         situation="행사·모임에서 여러 명에게 하나씩 돌리거나, 부담 없이 마음만 가볍게 전할 때",
         reason="1만원 안쪽 가격에 투명 박스까지 포함이라 답례품이나 단체 선물로 쓰기 좋아요."),

    dict(id="kinderjoy-doll-bouquet", name="킨더조이 레보니 인형꽃다발", cat="인형꽃다발",
         match=dict(all=["킨더조이"]),
         recipient=["아이"], occasion=["발표회·재롱잔치", "졸업·입학", "생일"],
         giftType=["아이용", "휴대전달형"],
         situation="유치원 재롱잔치·발표회·졸업식에서 아이에게 안겨줄 꽃다발이 필요할 때",
         reason="꽃과 인형에 킨더조이까지 들어 있어 아이가 바로 좋아하고, 사진에서 아이 얼굴을 가리지 않는 크기예요."),

    dict(id="character-doll-bouquet", name="캐릭터 인형꽃다발", cat="인형꽃다발",
         match=dict(any_of=["마이멜로디", "시나모롤"]),
         recipient=["아이"], occasion=["생일", "발표회·재롱잔치", "졸업·입학"],
         giftType=["아이용", "휴대전달형"],
         situation="산리오 캐릭터를 좋아하는 아이에게 생일·발표회 선물을 할 때",
         reason="마이멜로디·시나모롤 인형이 함께 있어 꽃다발을 받은 뒤에도 인형은 계속 곁에 둘 수 있어요."),

    dict(id="rebony-trophy-doll-bouquet", name="레보니 인형꽃다발 (트로피포트)", cat="인형꽃다발",
         match=dict(any_of=["깡총커플", "앙증맞은 레보니"]),
         recipient=["아이"], occasion=["발표회·재롱잔치", "졸업·입학"],
         giftType=["아이용"],
         situation="발표회·졸업식에서 아이를 주인공처럼 축하해주고 싶을 때",
         reason="트로피 모양 포트에 담겨 있어 시상식 분위기가 나고, 행사 뒤엔 방에 두는 소품이 돼요."),

    dict(id="glassdome-moodlamp", name="유리돔 무드등", cat="무드등",
         match=dict(any_of=["안개꽃핑크 무드등", "안개꽃파랑 무드등", "스타티스 무드등",
                            "핑크장미 무드등", "레드장미 무드등", "파랑장미 무드등"]),
         recipient=["연인", "배우자", "본인"],
         occasion=["기념일", "집들이·개업", "생일",
                   "화이트데이", "발렌타인데이", "크리스마스", "병문안"],
         giftType=["인테리어형"],
         situation="선물한 뒤에도 방에 두고 매일 보게 되는 것을 찾을 때, 또는 물·흙 없이 병문안에 가져갈 꽃을 찾을 때",
         reason="유리돔 안에 꽃을 담고 조명을 넣어, 꽃 선물이면서 동시에 인테리어 조명 역할을 해요. 물이 없어 병실에 가져가도 관리 부담이 없어요."),

    dict(id="classic-money-box", name="클래식투명 용돈박스", cat="용돈박스",
         match=dict(all=["클래식투명용돈박스"]),
         recipient=["부모님", "조부모님"], occasion=["생신·환갑", "어버이날", "설날·명절"],
         giftType=["현금동봉형"],
         situation="현금을 드리되 봉투 대신 형태를 갖춰 전하고 싶을 때",
         reason="투명 박스라 안에 든 꽃과 현금이 함께 보이고, 돈티슈·머니캡 두 방식 중 고를 수 있어요."),

    dict(id="premium-window-money-box", name="프리미엄창문 용돈박스", cat="용돈박스",
         match=dict(all=["프리미엄창문"]),
         recipient=["부모님", "조부모님"], occasion=["생신·환갑", "어버이날", "설날·명절"],
         giftType=["현금동봉형"],
         situation="용돈박스 중에서도 조금 더 격식 있는 구성을 원할 때",
         reason="창문형 박스에 꽃을 배치해 정면에서 봤을 때 정돈된 느낌이 나요. 스몰·라지 두 크기가 있습니다."),

    dict(id="sunflower-frame", name="해바라기 액자", cat="액자",
         match=dict(all=["해바라기 액자"]),
         recipient=["직장동료", "부모님", "본인"], occasion=["집들이·개업", "승진·퇴임"],
         giftType=["인테리어형"],
         situation="개업·집들이 선물로 벽에 걸 수 있는 것을 찾을 때",
         reason="재물운을 상징하는 해바라기를 액자에 담아, 매장이나 거실에 걸어두기 좋아요."),

    dict(id="cup-moodlamp-single", name="프리저브드 꽃 무드등 한송이", cat="무드등",
         match=dict(all=["프리저브드 꽃 무드등 한송이"]),
         recipient=["연인", "친구", "본인"],
         occasion=["생일", "기념일", "화이트데이", "발렌타인데이", "병문안"],
         giftType=["인테리어형"],
         situation="2만원대에서 오래 두고 볼 수 있는 선물을 찾을 때, 또는 병문안처럼 관리 부담 없는 꽃이 필요할 때",
         reason="컵 모양 무드등에 꽃 한 송이를 담은 구성이라 책상이나 머리맡에 두기 좋은 크기예요. 물을 안 갈아줘도 돼서 병실에도 부담 없어요."),

    dict(id="ionantha-moodlamp", name="이오난사 스칸디아모스 무드등", cat="무드등",
         match=dict(all=["이오난사"]),
         recipient=["직장동료", "본인", "친구"], occasion=["집들이·개업", "승진·퇴임"],
         giftType=["인테리어형"],
         situation="사무실 책상이나 작업 공간에 둘 식물 소품을 찾을 때",
         reason="이오난사와 스칸디아모스는 물 주기가 거의 필요 없는 공기정화 식물이라, 관리할 시간이 없어도 괜찮아요."),

    dict(id="centerpiece", name="센터피스", cat="센터피스",
         match=dict(all=["센터피스"]),
         recipient=["부모님", "본인", "직장동료"], occasion=["집들이·개업", "생신·환갑", "크리스마스"],
         giftType=["인테리어형"],
         situation="식탁이나 거실 테이블 가운데 놓을 장식을 찾을 때",
         reason="사방 어느 각도에서 봐도 균형이 잡히도록 만든 구성이라 테이블 가운데 두기에 맞아요."),

    dict(id="topiary-moodlamp", name="토피어리 수국 무드등", cat="무드등",
         match=dict(all=["토피어리"]),
         recipient=["연인", "부모님", "본인"], occasion=["집들이·개업", "생일", "기념일", "크리스마스"],
         giftType=["인테리어형"],
         situation="화분 형태의 조명 소품을 찾을 때",
         reason="동그랗게 다듬은 수국에 LED를 넣어, 불을 켜면 화분이 조명이 되는 구조예요."),

    dict(id="money-cake", name="용돈케이크 머니박스", cat="용돈박스",
         match=dict(all=["용돈케이크"]),
         recipient=["부모님", "조부모님"], occasion=["생신·환갑", "칠순·팔순", "어버이날"],
         giftType=["현금동봉형"],
         situation="현금만 드리기엔 민망하고 거창한 이벤트는 부담스러울 때",
         reason="케이크 모양에 현금을 꽂고 LED 조명까지 들어가, 촛불 대신 불을 켜고 축하하는 자리를 만들 수 있어요."),

    dict(id="hydrangea-moodlamp", name="수국꽃무드등", cat="무드등",
         match=dict(all=["수국꽃무드등"]),
         recipient=["연인", "배우자", "부모님"],
         occasion=["기념일", "프러포즈", "생신·환갑", "화이트데이", "발렌타인데이", "크리스마스"],
         giftType=["인테리어형"],
         situation="가장 공들인 선물 하나를 준비할 때",
         reason="프리저브드 수국을 가득 채우고 조명을 넣은 대표 상품이에요. 전원은 USB와 건전지 중 고를 수 있고, USB 쪽 선호가 조금 더 많습니다."),

    dict(id="flower-postcard", name="꽃엽서카드", cat="소품",
         match=dict(all=["꽃엽서카드"]),
         recipient=["연인", "부모님", "친구", "선생님", "직장동료", "아이"],
         occasion=["소소한 감사", "생일", "기념일",
                   "화이트데이", "발렌타인데이", "빼빼로데이", "크리스마스", "병문안"],
         giftType=["휴대전달형", "함께보내기"],
         situation="다른 선물에 손편지를 곁들이거나, 카드 한 장으로 마음만 전할 때",
         reason="크래프트 카드에 미니 드라이플라워를 붙여, 카드 자체가 작은 꽃다발이 돼요."),
]

# 2026-08-04부터 스마트스토어 링크는 xlsx의 "사이트링크" 열(E)에서 직접 읽는다 —
# 라인 안 아무 SKU 행에나 하나 채워두면 그 라인의 url이 된다(대표 1개 행이면 충분,
# 색상·사이즈는 같은 페이지의 옵션이므로 SKU마다 다 채울 필요 없음).
# URL_MAP은 xlsx에 없는 라인(예: 플라워클래스처럼 EXTRA_LINES에만 있는 것)을 위한
# 수동 폴백이다 — xlsx 쪽에 값이 있으면 그게 항상 우선한다.
URL_MAP = {
    "flower-class": "https://smartstore.naver.com/itecha/products/10537190482",
}

# xlsx에 없지만 실제 운영 중인 상품 (script-guide.md 기준)
EXTRA_LINES = [
    dict(id="flower-class", name="플라워클래스 (원데이)", cat="체험",
         recipient=["연인", "친구", "본인"], occasion=["데이트", "취미·자기계발", "기념일"],
         giftType=["체험형"],
         situation="물건 대신 함께 시간을 보내는 경험을 선물하고 싶을 때",
         reason="조소를 전공한 플로리스트가 직접 지도하고, 완성한 작품은 그대로 가져가실 수 있어요.",
         note="가격리스트(xlsx)에 없는 항목 — script-guide.md 3-8 기준으로 추가",
         variants=[
             dict(sku="원데이 클래스 70분 (1인)", price=50000, option="1인"),
             dict(sku="원데이 클래스 70분 (4인 이상, 1인당)", price=45000, option="4인 이상"),
         ]),
]

# ---------------------------------------------------------------------------
COLORS = ['블렉앤화이트', '블랙앤화이트', '트렌치바이올렛', '오리지날레드', '피치코랄',
          '라벤더퍼플', '소프트핑크', '스카이블루', '펄화이트', '황금옐로', '핑크핑크',
          '베이지&블랙', '화이트&핑크', '맑은배경', '황금배경', '블루자갈', '핑크자갈',
          '화이트자갈', '딥레드', '파랑', '보라', '레드', '핑크', '화이트', '블루',
          '노랑', '피치', '샤넬']
SIZES = ['중대형', '대형', '미니', '라지', '스몰', 'A4', 'A3', '디럭스',
         '5송이', '7송이', '10송이', '15송이', '20송이', '23송이', '25송이',
         '한송이', '두송이']
# 단 수는 사이즈와 별개 축(용돈케이크는 1단/2단이 가격을 가르는 실제 변수)
TIERS = ['1단', '2단']
OPTS = ['건전지방식', 'USB방식', '돈티슈', '머니캡']


def extract(name, pool):
    return next((t for t in pool if t in name), None)


def matches(name, m):
    if 'all' in m and not all(t in name for t in m['all']):
        return False
    if 'any_of' in m and not any(t in name for t in m['any_of']):
        return False
    if 'none' in m and any(t in name for t in m['none']):
        return False
    return True


def price_band(p):
    if p < 15000: return "1만원대"
    if p < 25000: return "2만원대"
    if p < 35000: return "3만원대"
    if p < 45000: return "4만원대"
    return "5만원 이상"


URL_RE = re.compile(r'^https?://')


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['Sheet1']
    skus = []
    bad_urls = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5, values_only=True):
        if not r[0]:
            continue
        url = str(r[4]).strip() if len(r) > 4 and r[4] else None
        if url and not URL_RE.match(url):
            bad_urls.append((r[0], url))
            url = None
        skus.append(dict(name=str(r[0]).strip(), price=r[1], mat=str(r[2] or '').strip(), url=url))

    buckets = {ln['id']: [] for ln in LINES}
    unmatched = []
    for s in skus:
        hit = next((ln for ln in LINES if matches(s['name'], ln['match'])), None)
        if hit:
            buckets[hit['id']].append(s)
        else:
            unmatched.append(s['name'])

    lines_out = []
    for ln in LINES:
        group = buckets[ln['id']]
        if not group:
            print(f"!! 빈 라인: {ln['name']}")
            continue
        prices = [g['price'] for g in group]
        mats = sorted({g['mat'] for g in group})
        urls_in_line = sorted({g['url'] for g in group if g['url']})
        if len(urls_in_line) > 1:
            print(f"?? 라인 '{ln['name']}' 안에 서로 다른 링크가 {len(urls_in_line)}개 있음 — 첫 번째만 사용:")
            for u in urls_in_line:
                print(f"     {u}")
        line_url = urls_in_line[0] if urls_in_line else URL_MAP.get(ln['id'])
        variants = []
        for g in group:
            v = OrderedDict(sku=g['name'], price=g['price'])
            for key, pool in (('color', COLORS), ('size', SIZES), ('tier', TIERS), ('option', OPTS)):
                val = extract(g['name'], pool)
                if val:
                    v[key] = val
            variants.append(v)
        lines_out.append(OrderedDict([
            ("id", ln['id']), ("name", ln['name']), ("category", ln['cat']),
            ("material", mats[0] if len(mats) == 1 else mats),
            ("priceMin", min(prices)), ("priceMax", max(prices)),
            ("priceBands", sorted({price_band(p) for p in prices},
                                  key=lambda b: ["1만원대", "2만원대", "3만원대", "4만원대", "5만원 이상"].index(b))),
            ("recipient", ln['recipient']), ("occasion", ln['occasion']),
            ("giftType", ln['giftType']),
            ("situation", ln['situation']), ("reason", ln['reason']),
            ("url", line_url),
            ("skuCount", len(group)), ("variants", variants),
        ]))

    for ln in EXTRA_LINES:
        prices = [v['price'] for v in ln['variants']]
        lines_out.append(OrderedDict([
            ("id", ln['id']), ("name", ln['name']), ("category", ln['cat']),
            ("material", "체험 프로그램"),
            ("priceMin", min(prices)), ("priceMax", max(prices)),
            ("priceBands", sorted({price_band(p) for p in prices},
                                  key=lambda b: ["1만원대", "2만원대", "3만원대", "4만원대", "5만원 이상"].index(b))),
            ("recipient", ln['recipient']), ("occasion", ln['occasion']),
            ("giftType", ln['giftType']),
            ("situation", ln['situation']), ("reason", ln['reason']),
            ("url", URL_MAP.get(ln['id'])),
            ("note", ln['note']),
            ("skuCount", len(ln['variants'])), ("variants", ln['variants']),
        ]))

    recipients, occasions, gift_types = [], [], []
    for l in lines_out:
        for v in l['recipient']:
            if v not in recipients: recipients.append(v)
        for v in l['occasion']:
            if v not in occasions: occasions.append(v)
        for v in l['giftType']:
            if v not in gift_types: gift_types.append(v)

    doc = OrderedDict([
        ("meta", OrderedDict([
            ("updated", "2026-08-04"),
            ("source", "테차상품가격리스트.xlsx"),
            ("skuCount", len(skus)),
            ("lineCount", len(lines_out)),
            ("note", "제품 라인 단위로 추천 태그를 관리한다. 색상·사이즈·전원방식 등 변형은 variants에 담긴다. "
                     "가격은 원(KRW). 태그를 고칠 때는 이 JSON을 고치고, xlsx는 가격 원본으로만 쓴다."),
        ])),
        ("axes", OrderedDict([
            ("recipient", recipients),
            ("occasion", occasions),
            ("giftType", gift_types),
            ("priceBand", ["1만원대", "2만원대", "3만원대", "4만원대", "5만원 이상"]),
        ])),
        ("lines", lines_out),
    ])

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    covered = sum(l['skuCount'] for l in lines_out if l['id'] != 'flower-class')
    print(f"라인 {len(lines_out)}개 / SKU {covered}개 매칭 (원본 {len(skus)}개)")
    if bad_urls:
        print(f"\n!! 유효하지 않은 링크 {len(bad_urls)}개 (http로 시작 안 함 — 무시하고 진행):")
        for name, url in bad_urls:
            print(f"   - {name}: '{url}'")
    no_url = [l['name'] for l in lines_out if not l.get('url')]
    if no_url:
        print(f"\n-- 아직 링크 없는 라인 {len(no_url)}개: {', '.join(no_url)}")
    if unmatched:
        print(f"\n!! 미분류 {len(unmatched)}개:")
        for u in unmatched:
            print("   -", u)
    else:
        print("미분류 없음 — 147개 전부 매칭됨")
    print(f"\n저장: {OUT}")
    for l in lines_out:
        print(f"  {l['skuCount']:3d}종 {l['priceMin']:>6,}~{l['priceMax']:>6,}원  {l['name']}")


main()
